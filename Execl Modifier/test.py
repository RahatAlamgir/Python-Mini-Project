from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_marksheet(sheet, title="Marksheet"):
    # Title
    sheet.merge_cells("A1:F1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Roll", "Name", "Subject 1", "Subject 2", "Subject 3", "Total"]

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Table rows (empty)
    for row in range(4, 14):  # 10 students
        for col in range(1, 7):
            cell = sheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    # Borders
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(3, 14):
        for col in range(1, 7):
            sheet.cell(row=row, column=col).border = border

    # Column width
    widths = [10, 20, 15, 15, 15, 15]
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + i)].width = w


def create_excel(file_name="marksheet.xlsx", num_sheets=3):
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create multiple sheets
    for i in range(num_sheets):
        sheet_name = f"Class_{i+1}"
        sheet = wb.create_sheet(title=sheet_name)
        create_marksheet(sheet, title=f"{sheet_name} Marksheet")

    wb.save(file_name)
    print(f"✅ Excel file '{file_name}' created with {num_sheets} sheets.")


# Run
create_excel("marksheet.xlsx", num_sheets=3)